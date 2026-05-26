import os
import json
import time
import logging
import re
import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from datetime import datetime, timezone
from typing import Any, Optional

from dotenv import load_dotenv
from azure.core.credentials import AzureKeyCredential
from azure.core.exceptions import HttpResponseError, ServiceRequestError
from azure.ai.documentintelligence import DocumentIntelligenceClient
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

load_dotenv()

AZURE_ENDPOINT = os.environ["AZURE_DI_ENDPOINT"]
AZURE_KEY = os.environ["AZURE_DI_KEY"]
MODEL_ID = os.environ["AZURE_DI_MODEL_ID"]

INPUT_DIR = os.getenv("DOCAI_INPUT_DIR", "./input")
INTEL_OUTPUT_DIR = os.getenv("DOCAI_INTEL_OUTPUT_DIR", "./intel_output")
CLEAN_LABELED_DIR = os.getenv("DOCAI_CLEAN_LABELED_DIR", "./clean_labeled_data")
EXCEL_OUTPUT_DIR = os.getenv("DOCAI_EXCEL_OUTPUT_DIR", "./excel_output")
EXCEL_FILENAME = os.getenv("DOCAI_EXCEL_FILENAME", "licenses.xlsx")
RUNS_DIR = os.getenv("DOCAI_RUNS_DIR", "./runs")

MAX_WORKERS = int(os.getenv("DOCAI_MAX_WORKERS", "4"))

# Identifier field used to join array sheets back to main rows.
# Set DOCAI_ID_FIELD to the exact label in your model (e.g. "License Number").
ID_FIELD = os.getenv("DOCAI_ID_FIELD", "License Number")

MAX_FILE_SIZE_BYTES = int(os.getenv("DOCAI_MAX_FILE_SIZE_MB", "500")) * 1024 * 1024
MAX_RETRIES = int(os.getenv("DOCAI_MAX_RETRIES", "3"))

# Columns pinned to the front of the main sheet, in this order.
PRIORITY_COLUMNS = [
    "source_file",
    "document_type",
    "document_confidence",
    "id_value",
    "page_count",
    "documents_detected",
    "field_count",
    "processed_at",
]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)


def normalize_field_name(name: str) -> str:
    """Lowercase, underscore-separated, ascii-safe field name."""
    cleaned = re.sub(r"[^a-zA-Z0-9]+", "_", name).strip("_").lower()
    return cleaned or "unknown_field"


# STAGE 1 — Send PDFs to Azure Document Intelligence, save raw responses
def get_di_client() -> DocumentIntelligenceClient:
    return DocumentIntelligenceClient(
        endpoint=AZURE_ENDPOINT,
        credential=AzureKeyCredential(AZURE_KEY),
    )


def analyze_pdf(client: DocumentIntelligenceClient,
                file_path: Path) -> Optional[dict]:
    file_size = file_path.stat().st_size
    if file_size > MAX_FILE_SIZE_BYTES:
        log.warning(
            "Skipping %s — %.0f MB exceeds the configured size limit.",
            file_path.name, file_size / (1024 * 1024),
        )
        return None

    log.info("Analyzing: %s (%.1f MB)", file_path.name, file_size / (1024 * 1024))

    with open(file_path, "rb") as f:
        file_bytes = f.read()

    last_err = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            poller = client.begin_analyze_document(
                model_id=MODEL_ID,
                body=file_bytes,
                content_type="application/pdf",
            )
            return poller.result().as_dict()
        except (HttpResponseError, ServiceRequestError) as e:
            last_err = e
            status = getattr(e, "status_code", None)
            # Retry on throttling and transient server errors only.
            if status not in (429, 500, 502, 503, 504) and not isinstance(e, ServiceRequestError):
                raise
            backoff = 2 ** attempt
            log.warning("  Attempt %d/%d failed (%s) — retrying in %ds",
                        attempt, MAX_RETRIES, status or type(e).__name__, backoff)
            time.sleep(backoff)

    raise last_err


def save_raw_output(result: dict, source_filename: str, output_dir: Path) -> Path:
    out_name = Path(source_filename).stem + "_raw.json"
    out_path = output_dir / out_name

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, default=str)

    log.info("  → Raw output saved: %s", out_path.name)
    return out_path


def _process_one_pdf(client: DocumentIntelligenceClient,
                     file_path: Path,
                     output_path: Path,
                     force: bool) -> dict:
    """Process a single PDF. Returns a manifest record."""
    record: dict[str, Any] = {
        "source": file_path.name,
        "size_bytes": file_path.stat().st_size,
        "status": "ok",
        "raw_path": None,
        "duration_s": 0.0,
        "error": None,
    }

    expected_raw = output_path / (file_path.stem + "_raw.json")
    if expected_raw.exists() and not force:
        log.info("Skipping %s — raw output already exists (use --force to re-run)",
                 file_path.name)
        record["status"] = "skipped"
        record["raw_path"] = str(expected_raw)
        return record

    started = time.monotonic()
    try:
        result = analyze_pdf(client, file_path)
        if result is None:
            record["status"] = "too_large"
        else:
            rp = save_raw_output(result, file_path.name, output_path)
            record["raw_path"] = str(rp)
    except Exception as e:
        log.exception("Failed to analyze %s", file_path.name)
        record["status"] = "failed"
        record["error"] = f"{type(e).__name__}: {e}"
    finally:
        record["duration_s"] = round(time.monotonic() - started, 2)

    return record


def run_stage_1(force: bool = False,
                manifest: Optional[dict] = None) -> list[Path]:
    input_path = Path(INPUT_DIR)
    output_path = Path(INTEL_OUTPUT_DIR)
    output_path.mkdir(parents=True, exist_ok=True)

    files = sorted(
        f for f in input_path.iterdir()
        if f.is_file() and f.suffix.lower() == ".pdf"
    )

    if not files:
        log.warning("No PDF files found in %s", input_path)
        if manifest is not None:
            manifest["stage_1"] = {"files": [], "workers": MAX_WORKERS}
        return []

    log.info("Found %d PDF(s) in %s — analyzing with %d worker(s)",
             len(files), input_path, MAX_WORKERS)

    client = get_di_client()
    records: list[dict] = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        future_to_file = {
            pool.submit(_process_one_pdf, client, fp, output_path, force): fp
            for fp in files
        }
        for fut in as_completed(future_to_file):
            records.append(fut.result())

    records.sort(key=lambda r: r["source"])

    raw_paths = [Path(r["raw_path"]) for r in records if r["raw_path"]]
    counts = {"ok": 0, "skipped": 0, "failed": 0, "too_large": 0}
    for r in records:
        counts[r["status"]] = counts.get(r["status"], 0) + 1

    log.info("Stage 1 complete — ok=%d skipped=%d failed=%d too_large=%d",
             counts["ok"], counts["skipped"], counts["failed"], counts["too_large"])

    if manifest is not None:
        manifest["stage_1"] = {
            "workers": MAX_WORKERS,
            "counts": counts,
            "files": records,
        }

    return raw_paths


# STAGE 2 — Clean raw JSON into human-readable, labeled JSON
def extract_field_value(field: dict) -> dict:
    field_type = field.get("type", "string")
    content = field.get("content", "")
    confidence = field.get("confidence", 0.0)
    value = field.get("value")

    extracted = {
        "type": field_type,
        "content": content.strip() if content else "",
        "confidence": round(confidence, 4) if confidence else 0.0,
    }

    if field_type in ("string", "phoneNumber", "countryRegion") and isinstance(value, str):
        extracted["value"] = value
    elif field_type in ("number", "integer") and value is not None:
        extracted["value"] = value
    elif field_type in ("date", "time") and value is not None:
        extracted["value"] = str(value)
    elif field_type == "selectionMark":
        extracted["value"] = field.get("value", "")
    elif field_type == "currency" and isinstance(value, dict):
        extracted["value"] = value.get("amount")
        extracted["currency_code"] = value.get("currencyCode", "")
        extracted["currency_symbol"] = value.get("currencySymbol", "")
    elif field_type == "address" and isinstance(value, dict):
        addr_parts = []
        for k in ["streetAddress", "city", "state", "postalCode", "countryRegion"]:
            if value.get(k):
                addr_parts.append(value[k])
        extracted["value"] = ", ".join(addr_parts) if addr_parts else content
        extracted["address_components"] = {k: v for k, v in value.items() if v}
    elif field_type == "array":
        value_array = field.get("valueArray", [])
        extracted["value"] = f"[{len(value_array)} items]"
        extracted["items"] = []
        for i, item in enumerate(value_array):
            if not isinstance(item, dict):
                continue
            item_fields = {}
            if item.get("type") == "object":
                obj = item.get("valueObject", {})
                for sub_name, sub_field in obj.items():
                    if isinstance(sub_field, dict):
                        item_fields[sub_name] = extract_field_value(sub_field)
            else:
                item_fields["value"] = item.get("content")
            extracted["items"].append({"index": i, "fields": item_fields})
    elif field_type == "object" and isinstance(value, dict):
        extracted["value"] = content
        extracted["sub_fields"] = {}
        for sub_name, sub_field in value.items():
            if isinstance(sub_field, dict):
                extracted["sub_fields"][sub_name] = extract_field_value(sub_field)
    else:
        extracted["value"] = value if value is not None else content

    return extracted


def clean_raw_file(raw_path: Path, output_dir: Path) -> Path:
    with open(raw_path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    documents = raw.get("documents", [])
    pages = raw.get("pages", [])

    all_doc_results = []
    for doc in documents:
        doc_type = doc.get("docType", "unknown")
        confidence = doc.get("confidence", 0.0)
        fields = doc.get("fields", {})

        cleaned_fields = {}
        for field_name, field_data in fields.items():
            if isinstance(field_data, dict):
                cleaned_fields[field_name] = {
                    "label": field_name,
                    "label_normalized": normalize_field_name(field_name),
                    **extract_field_value(field_data),
                }

        all_doc_results.append({
            "document_type": doc_type,
            "document_confidence": round(confidence, 4),
            "field_count": len(cleaned_fields),
            "fields": cleaned_fields,
        })

    all_labels = set()
    for dr in all_doc_results:
        all_labels.update(dr["fields"].keys())

    cleaned = {
        "source_file": raw_path.stem.replace("_raw", ""),
        "processed_at": datetime.now(timezone.utc).isoformat(),
        "page_count": len(pages),
        "documents_detected": len(documents),
        "labels_found": sorted(all_labels),
        "documents": all_doc_results,
    }

    full_text = raw.get("content", "")
    cleaned["full_text_preview"] = (
        full_text[:2000] + "…" if len(full_text) > 2000 else full_text
    )

    out_name = raw_path.stem.replace("_raw", "_clean") + ".json"
    out_path = output_dir / out_name
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(cleaned, f, indent=2, ensure_ascii=False)

    log.info("  → Clean output: %s  (%d document(s), %d label(s))",
             out_name, len(documents), len(all_labels))
    return out_path


def run_stage_2(raw_paths: Optional[list[Path]] = None,
                manifest: Optional[dict] = None) -> list[Path]:
    intel_path = Path(INTEL_OUTPUT_DIR)
    clean_path = Path(CLEAN_LABELED_DIR)
    clean_path.mkdir(parents=True, exist_ok=True)

    if raw_paths is None:
        raw_paths = sorted(intel_path.glob("*_raw.json"))

    if not raw_paths:
        log.warning("No raw output files found in %s", intel_path)
        if manifest is not None:
            manifest["stage_2"] = {"files": []}
        return []

    log.info("Cleaning %d raw file(s)…", len(raw_paths))
    clean_paths = []
    records: list[dict] = []
    for rp in raw_paths:
        rec: dict[str, Any] = {"source": rp.name, "status": "ok",
                               "clean_path": None, "error": None}
        try:
            cp = clean_raw_file(rp, clean_path)
            clean_paths.append(cp)
            rec["clean_path"] = str(cp)
        except Exception as e:
            log.exception("Failed to clean %s", rp.name)
            rec["status"] = "failed"
            rec["error"] = f"{type(e).__name__}: {e}"
        records.append(rec)

    log.info("Stage 2 complete — %d clean file(s) written.", len(clean_paths))

    if manifest is not None:
        manifest["stage_2"] = {"files": records}

    return clean_paths


# STAGE 3 — Aggregate clean JSON into an Excel workbook
def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(column)].width = min(max_length + 2, 80)


def _order_columns(columns: set[str], priority: list[str]) -> list[str]:
    """Priority columns first (in given order), then the rest alphabetically."""
    front = [c for c in priority if c in columns]
    rest = sorted(c for c in columns if c not in front)
    return front + rest


def _write_sheet(ws, rows: list[dict], priority: list[str]) -> None:
    columns = _order_columns({k for r in rows for k in r.keys()}, priority)
    for col_num, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_num, value=col_name)
    for row_num, row_data in enumerate(rows, 2):
        for col_num, col_name in enumerate(columns, 1):
            ws.cell(row=row_num, column=col_num, value=row_data.get(col_name))
    autosize_columns(ws)


def run_stage_3(clean_paths: Optional[list[Path]] = None,
                manifest: Optional[dict] = None) -> Optional[Path]:
    clean_dir = Path(CLEAN_LABELED_DIR)
    excel_dir = Path(EXCEL_OUTPUT_DIR)
    excel_dir.mkdir(parents=True, exist_ok=True)

    if clean_paths is None:
        clean_paths = sorted(clean_dir.glob("*_clean.json"))

    if not clean_paths:
        log.warning("No clean files found in %s", clean_dir)
        return None

    main_rows: list[dict] = []
    array_tables: dict[str, list[dict]] = {}

    for json_file in clean_paths:
        log.info("Aggregating: %s", json_file.name)
        with open(json_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        source_file = data.get("source_file")

        for doc in data.get("documents", []):
            fields = doc.get("fields", {})

            id_value = None
            if ID_FIELD in fields:
                id_value = fields[ID_FIELD].get("value")

            row = {
                "source_file": source_file,
                "processed_at": data.get("processed_at"),
                "page_count": data.get("page_count"),
                "documents_detected": data.get("documents_detected"),
                "document_type": doc.get("document_type"),
                "document_confidence": doc.get("document_confidence"),
                "field_count": doc.get("field_count"),
                "id_value": id_value,
            }

            for field_name, field_data in fields.items():
                field_type = field_data.get("type")
                normalized = field_data.get(
                    "label_normalized", normalize_field_name(field_name)
                )

                if field_type != "array":
                    row[normalized] = field_data.get("value")
                    conf = field_data.get("confidence")
                    if conf is not None:
                        row[f"{normalized}__conf"] = conf
                else:
                    array_tables.setdefault(normalized, [])
                    for idx, item in enumerate(field_data.get("items", [])):
                        item_row = {
                            "source_file": source_file,
                            "id_value": id_value,
                            "item_index": idx,
                        }
                        for sub_name, sub_data in item.get("fields", {}).items():
                            col = normalize_field_name(sub_name)
                            item_row[col] = sub_data.get("value")
                            conf = sub_data.get("confidence")
                            if conf is not None:
                                item_row[f"{col}__conf"] = conf
                        array_tables[normalized].append(item_row)

            main_rows.append(row)

    wb = Workbook()
    wb.remove(wb.active)

    ws_main = wb.create_sheet("main")
    _write_sheet(ws_main, main_rows, PRIORITY_COLUMNS)

    array_priority = ["source_file", "id_value", "item_index"]
    for sheet_name, rows in array_tables.items():
        ws = wb.create_sheet(sheet_name[:31])
        _write_sheet(ws, rows, array_priority)

    out_path = excel_dir / EXCEL_FILENAME
    wb.save(out_path)

    log.info("Stage 3 complete — workbook: %s  (%d main row(s), %d array sheet(s))",
             out_path, len(main_rows), len(array_tables))
    log.info("  Joining on field: %r → column 'id_value'", ID_FIELD)

    if manifest is not None:
        manifest["stage_3"] = {
            "workbook": str(out_path),
            "main_rows": len(main_rows),
            "array_sheets": {name: len(rows) for name, rows in array_tables.items()},
            "id_field": ID_FIELD,
        }

    return out_path


def write_manifest(manifest: dict) -> Path:
    runs_dir = Path(RUNS_DIR)
    runs_dir.mkdir(parents=True, exist_ok=True)
    out_path = runs_dir / f"{manifest['run_id']}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, default=str)
    log.info("Run manifest: %s", out_path)
    return out_path


# MAIN
def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Azure Document Intelligence pipeline: PDF → raw JSON → clean JSON → Excel.",
    )
    p.add_argument(
        "--stage",
        choices=["1", "2", "3", "all"],
        default="all",
        help="Run a single stage or the whole pipeline (default: all).",
    )
    p.add_argument(
        "--force",
        action="store_true",
        help="Re-run Stage 1 even if raw output for a PDF already exists.",
    )
    return p.parse_args()


def main() -> None:
    args = parse_args()

    started_at = datetime.now(timezone.utc)
    run_id = started_at.strftime("%Y%m%dT%H%M%SZ")
    manifest: dict[str, Any] = {
        "run_id": run_id,
        "started_at": started_at.isoformat(),
        "finished_at": None,
        "stages_run": [],
        "args": {"stage": args.stage, "force": args.force},
        "config": {
            "model_id": MODEL_ID,
            "max_workers": MAX_WORKERS,
            "max_retries": MAX_RETRIES,
            "id_field": ID_FIELD,
        },
    }

    log.info("=" * 60)
    log.info("Azure Document Intelligence Pipeline — run_id=%s stage=%s force=%s",
             run_id, args.stage, args.force)
    log.info("=" * 60)

    raw_paths = None
    clean_paths = None

    try:
        if args.stage in ("1", "all"):
            input_path = Path(INPUT_DIR)
            if not input_path.is_dir():
                log.error("Input directory does not exist: %s", input_path)
                return
            log.info("-" * 40)
            log.info("STAGE 1: Azure Document Intelligence")
            log.info("-" * 40)
            manifest["stages_run"].append("1")
            raw_paths = run_stage_1(force=args.force, manifest=manifest)

        if args.stage in ("2", "all"):
            log.info("-" * 40)
            log.info("STAGE 2: Clean & Label")
            log.info("-" * 40)
            manifest["stages_run"].append("2")
            clean_paths = run_stage_2(raw_paths, manifest=manifest)

        if args.stage in ("3", "all"):
            log.info("-" * 40)
            log.info("STAGE 3: Aggregate to Excel")
            log.info("-" * 40)
            manifest["stages_run"].append("3")
            run_stage_3(clean_paths, manifest=manifest)
    finally:
        finished_at = datetime.now(timezone.utc)
        manifest["finished_at"] = finished_at.isoformat()
        manifest["duration_s"] = round(
            (finished_at - started_at).total_seconds(), 2
        )
        write_manifest(manifest)

    log.info("=" * 60)
    log.info("Pipeline complete!")
    log.info("  Raw output:    %s", INTEL_OUTPUT_DIR)
    log.info("  Clean labeled: %s", CLEAN_LABELED_DIR)
    log.info("  Excel output:  %s", EXCEL_OUTPUT_DIR)
    log.info("  Run manifest:  %s/", RUNS_DIR)
    log.info("=" * 60)


if __name__ == "__main__":
    main()
